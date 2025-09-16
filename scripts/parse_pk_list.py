import zipfile
import csv
import io
import os
import json
from pathlib import Path
from openpyxl import load_workbook


def detect_text(content: bytes) -> str:
    for enc in ["utf-8-sig", "utf-16", "cp1254", "iso-8859-9", "utf-8"]:
        try:
            return content.decode(enc)
        except Exception:
            continue
    # last resort
    return content.decode("latin-1", errors="ignore")


def normalize_string(s: str) -> str:
    return (s or "").strip()


def main():
    zip_path = Path("data/ptt/pk_list.zip")
    extract_dir = Path("data/ptt/pk_list")
    out_json = Path("data/ptt/pk_list_normalized.json")
    out_dir = Path("data/ptt/pk_out")

    extract_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(extract_dir)

    # detect file type
    csv_or_txt = None
    xlsx = None
    for p in extract_dir.rglob("*"):
        suf = p.suffix.lower()
        if suf in [".csv", ".txt", ".tsv"]:
            csv_or_txt = p
            break
        if suf in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            xlsx = p
    rows = []
    if csv_or_txt:
        raw = csv_or_txt.read_bytes()
        text = detect_text(raw)
        delimiter_candidates = [';', ',', '\t']
        best_delim = None
        for delim in delimiter_candidates:
            reader = csv.reader(io.StringIO(text), delimiter=("\t" if delim == "\t" else delim))
            test = list(reader)[:5]
            if test and max(len(r) for r in test) >= 3:
                best_delim = delim
                break
        if not best_delim:
            best_delim = ';'
        reader = csv.reader(io.StringIO(text), delimiter=("\t" if best_delim == "\t" else best_delim))
        rows = list(reader)
    elif xlsx:
        wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
    else:
        print("No parseable file found (CSV/TXT/XLSX)")
        return

    # Heuristic: expected columns contain city, district, neighborhood, postal code
    # Often header exists; try to detect header by checking non-numeric postal code col
    header = rows[0]
    # Guess column indices by simple heuristics
    # Find a column with 5-digit postal code
    postal_idx = None
    for i in range(len(header)):
        for r in rows[1:20]:
            if i < len(r):
                v = (r[i] or "").strip()
                if v.isdigit() and len(v) == 5:
                    postal_idx = i
                    break
        if postal_idx is not None:
            break

    # Assume first three columns are city, district, neighborhood if present
    # Adjust if postal at end
    city_idx, district_idx, neighborhood_idx = 0, 1, 2
    if postal_idx is not None and postal_idx < 3:
        # shift indices to avoid postal
        city_idx = 0 if postal_idx != 0 else 1
        district_idx = 1 if postal_idx != 1 else 2
        neighborhood_idx = 2 if postal_idx != 2 else 3

    # Skip header row if it contains letters like "IL" etc.
    start_row = 1
    header_join = " ".join(header).lower()
    if not any(k in header_join for k in ["il", "ilce", "mahalle", "posta", "kod"]):
        # likely no header
        start_row = 0

    cities = {}
    districts = set()
    neighborhoods = []

    for r in rows[start_row:]:
        if len(r) <= max(city_idx, district_idx, neighborhood_idx):
            continue
        city = normalize_string(r[city_idx])
        district = normalize_string(r[district_idx])
        neighborhood = normalize_string(r[neighborhood_idx])
        postal = normalize_string(r[postal_idx]) if postal_idx is not None and postal_idx < len(r) else ""
        if not city or not district or not neighborhood:
            continue
        cities[city] = True
        districts.add((city, district))
        neighborhoods.append({
            "city": city,
            "district": district,
            "neighborhood": neighborhood,
            "postal_code": postal
        })

    data = {
        "cities": sorted(list(cities.keys())),
        "districts": sorted([{"city": c, "district": d} for (c, d) in districts], key=lambda x: (x["city"], x["district"])),
        "neighborhoods": neighborhoods
    }

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    # also write split files for easy bulk import
    out_dir.mkdir(parents=True, exist_ok=True)
    Path(out_dir / 'cities.json').write_text(json.dumps(data['cities'], ensure_ascii=False), encoding='utf-8')
    Path(out_dir / 'districts.json').write_text(json.dumps(data['districts'], ensure_ascii=False), encoding='utf-8')
    # split neighborhoods
    batch_size = 2000
    nb = data['neighborhoods']
    for i in range(0, len(nb), batch_size):
        chunk = nb[i:i+batch_size]
        Path(out_dir / f'neighborhoods_{i//batch_size:04d}.json').write_text(json.dumps(chunk, ensure_ascii=False), encoding='utf-8')

    # NEW: group neighborhoods per (city, district)
    by_district_dir = out_dir / 'neighborhoods_by_district'
    by_district_dir.mkdir(parents=True, exist_ok=True)
    groups = {}
    for item in nb:
        key = f"{item['city']}__{item['district']}"
        groups.setdefault(key, []).append(item)
    for key, arr in groups.items():
        safe = key.replace('/', '-').replace('\\', '-').replace(' ', '_')
        out_path = by_district_dir / f"{safe}.json"
        out_path.write_text(json.dumps(arr, ensure_ascii=False, indent=2), encoding='utf-8')

    print("OK:", out_json, "cities:", len(data["cities"]), "districts:", len(data["districts"]), "neighborhoods:", len(data["neighborhoods"]))


if __name__ == "__main__":
    main()


